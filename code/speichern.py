import sys
sys.path.append("U:\locallib\openpyxl-2.3.4")
sys.path.append("U:\locallib\jdcal-1.2")
sys.path.append("U:\locallib\et_xmlfile-1.0.1")
from openpyxl import *
"""
öffnet ein Excel Workbook kann alle antworten auf einmal eintragen und speichern
"""
class Save:
    
    def __init__(self, filename):
        self.fname = filename
        self.wb = load_workbook(filename = filename)
        #self.ws = self.wb.get_active_sheet()
        self.ws = self.wb.active
        self.ws.title = "Test"
        self.row = self.ws.max_row
        self.antworten=[]
    
    def change_fname(self,fname):
        self.fname=fname
    
    #speichert alle antworten
    def save(self):
        self.write_all()
        print("speichern")
        self.wb.save("..\\Daten\\"+self.fname)
    
    #schreibt alle antworten
    def write_all(self):
        antworten = self.antworten
        row=self.row
        self.ws.cell(row=row, column=1, value="Person "+str(row))
        col = 2
        for antwort in antworten:
            for a in antwort:
                self.ws.cell(row=row, column=col, value=a)
                col+=1
        self.ws.cell(row=row+1, column=1, value="END")
    
    #antworten werden zum speichern appended
    def save_antworten(self,antworten):
        #print(antworten)
        self.antworten.append(antworten)
    
    def reset_antworten(self):
        self.antworten=[]